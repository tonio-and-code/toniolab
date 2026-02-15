import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

# 新しいワークブックを作成
wb = Workbook()
ws = wb.active
ws.title = "工事注文書"

# --- スタイル定義 ---
border_thin = Side(border_style="thin", color="000000")
border_double = Side(border_style="double", color="000000")
border_box = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
align_center = Alignment(horizontal="center", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")
align_left = Alignment(horizontal="left", vertical="center")
font_title = Font(name="游ゴシック", size=18, bold=True)
font_header = Font(name="游ゴシック", size=11, bold=True)
font_bold = Font(name="游ゴシック", size=11, bold=True)
font_normal = Font(name="游ゴシック", size=11)

# --- 列幅の設定 ---
ws.column_dimensions['A'].width = 5   # No
ws.column_dimensions['B'].width = 40  # 項目
ws.column_dimensions['C'].width = 10  # 数量
ws.column_dimensions['D'].width = 8   # 単位
ws.column_dimensions['E'].width = 12  # 単価
ws.column_dimensions['F'].width = 15  # 金額
ws.column_dimensions['G'].width = 20  # 備考

# --- ヘッダー部 ---
# タイトル
ws.merge_cells('A1:G1')
ws['A1'] = "注　文　書"
ws['A1'].font = font_title
ws['A1'].alignment = align_center

# 日付欄（右上一段下）
ws.merge_cells('F2:G2')
ws['F2'] = "発注日：202__年__月__日"
ws['F2'].font = font_normal
ws['F2'].alignment = align_right

# 宛先（左側）
ws.merge_cells('A4:C4')
ws['A4'] = "トップワン　石渡 友和　殿"
ws['A4'].font = Font(name="游ゴシック", size=14, bold=True, underline="single")
ws['A4'].alignment = Alignment(horizontal="left", vertical="bottom")

ws.merge_cells('A5:C5')
ws['A5'] = "下記の内容にて発注いたします。"
ws['A5'].font = font_normal
ws['A5'].alignment = align_left

# 発注者（右側）
ws.merge_cells('E4:G4')
ws['E4'] = "有限会社 イワサキ内装"
ws['E4'].font = font_bold
ws['E4'].alignment = align_left

ws.merge_cells('E5:G5')
ws['E5'] = "〒XXX-XXXX" # 住所プレースホルダー
ws['E5'].font = font_normal
ws['E5'].alignment = align_left

ws.merge_cells('E6:G6')
ws['E6'] = "住所：_____________________"
ws['E6'].font = font_normal
ws['E6'].alignment = align_left

ws.merge_cells('E7:G7')
ws['E7'] = "TEL：____________ / FAX：____________"
ws['E7'].font = font_normal
ws['E7'].alignment = align_left

# --- 工事概要（枠付き） ---
start_row = 9
labels = ["工事名称", "工事場所", "工期", "支払条件", "契約形態"]
values = ["", "有限会社イワサキ内装", "", "末締め　翌月末支払", "請負"]

for i, (label, val) in enumerate(zip(labels, values)):
    r = start_row + i
    ws[f'A{r}'] = label
    ws[f'A{r}'].font = font_header
    ws[f'A{r}'].alignment = align_center
    ws[f'A{r}'].border = border_box

    ws.merge_cells(f'B{r}:G{r}')
    ws[f'B{r}'] = val
    ws[f'B{r}'].font = font_normal
    ws[f'B{r}'].alignment = align_left
    ws[f'B{r}'].border = border_box

# --- 明細テーブル ---
table_head_row = 15
headers = ["No.", "工事項目・内容", "数量", "単位", "単価", "金額", "備考"]
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=table_head_row, column=col_idx)
    cell.value = header
    cell.font = font_header
    cell.alignment = align_center
    cell.border = border_box
    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# 行データの作成 (10行分)
for i in range(1, 11):
    r = table_head_row + i
    # No
    ws.cell(row=r, column=1, value=i).alignment = align_center
    # 内容
    ws.cell(row=r, column=2).alignment = align_left
    # 数量
    ws.cell(row=r, column=3).number_format = '#,##0'
    # 単位
    ws.cell(row=r, column=4).alignment = align_center
    # 単価
    ws.cell(row=r, column=5).number_format = '#,##0'
    # 金額 (計算式)
    ws.cell(row=r, column=6, value=f'=IF(AND(C{r}<>"",E{r}<>""), C{r}*E{r}, "")').number_format = '#,##0'
    # 備考
    ws.cell(row=r, column=7).alignment = align_left

    # 罫線適用
    for c in range(1, 8):
        ws.cell(row=r, column=c).border = border_box

# --- 合計欄 ---
last_row = table_head_row + 10
total_start_row = last_row + 1

# 小計
ws.merge_cells(f'A{total_start_row}:E{total_start_row}')
ws[f'A{total_start_row}'] = "小　計"
ws[f'A{total_start_row}'].alignment = align_center
ws[f'A{total_start_row}'].font = font_header
for c in range(1, 6): ws.cell(row=total_start_row, column=c).border = border_box

ws[f'F{total_start_row}'] = f'=SUM(F{table_head_row+1}:F{last_row})'
ws[f'F{total_start_row}'].number_format = '"¥"#,##0'
ws[f'F{total_start_row}'].font = font_bold
ws[f'F{total_start_row}'].border = border_box
ws[f'G{total_start_row}'].border = border_box # 備考の空セル

# 消費税
tax_row = total_start_row + 1
ws.merge_cells(f'A{tax_row}:E{tax_row}')
ws[f'A{tax_row}'] = "消費税（10%）"
ws[f'A{tax_row}'].alignment = align_center
ws[f'A{tax_row}'].font = font_header
for c in range(1, 6): ws.cell(row=tax_row, column=c).border = border_box

ws[f'F{tax_row}'] = f'=INT(F{total_start_row}*0.1)'
ws[f'F{tax_row}'].number_format = '"¥"#,##0'
ws[f'F{tax_row}'].font = font_bold
ws[f'F{tax_row}'].border = border_box
ws[f'G{tax_row}'].border = border_box

# 合計（税込）
grand_row = tax_row + 1
ws.merge_cells(f'A{grand_row}:E{grand_row}')
ws[f'A{grand_row}'] = "合計金額（税込）"
ws[f'A{grand_row}'].alignment = align_center
ws[f'A{grand_row}'].font = Font(name="游ゴシック", size=12, bold=True)
for c in range(1, 6): ws.cell(row=grand_row, column=c).border = border_box

ws[f'F{grand_row}'] = f'=F{total_start_row}+F{tax_row}'
ws[f'F{grand_row}'].number_format = '"¥"#,##0'
ws[f'F{grand_row}'].font = Font(name="游ゴシック", size=12, bold=True)
ws[f'F{grand_row}'].border = border_box
ws[f'G{grand_row}'].border = border_box

# --- 備考欄 ---
note_row = grand_row + 2
ws[f'A{note_row}'] = "【備　考】"
ws[f'A{note_row}'].font = font_bold
ws.merge_cells(f'A{note_row+1}:G{note_row+5}')
ws[f'A{note_row+1}'].border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)
ws[f'A{note_row+1}'].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 保存先ディレクトリの作成
output_dir = os.path.join(os.path.dirname(__file__), "..", "output")
os.makedirs(output_dir, exist_ok=True)

# ファイル名にタイムスタンプを追加
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
file_path = os.path.join(output_dir, f"工事注文書_{timestamp}.xlsx")

# 保存
wb.save(file_path)
print(f"注文書を生成しました: {file_path}")
